from bs4 import BeautifulSoup
import requests, sys
from openpyxl import load_workbook

############################################
if len(sys.argv) == 5:
    parentPath = str(sys.argv[1])+'\\'
    fileToProcess = str(sys.argv[2])+'.xlsx'
    refNoCol = int(sys.argv[3])
    remarksCol = int(sys.argv[4])
else:
######## Section for Manual Input ##########
    parentPath = input('Enter folder path (e.g.,D:\\Lesco\\Recovery\\01-23) :')+'\\'
    fileToProcess = input('Enter file name (e.g.,WorkingBookDT25-10k) :')+'.xlsx'
    refNoCol = int(input('Enter reference number column number (e.g.,3) :'))
    remarksCol = int(input('Enter remarks number (e.g.,3) :'))
isRefComplete = True
batchCol = 2
subDiv  =   "11216"
############################################

wb = load_workbook(filename = parentPath+fileToProcess)

for sheet in wb.sheetnames:
    ws = wb[sheet]
    print("Testing of "+ sheet +"started from row 2 to "+str(ws.max_row+1))
    for row in range(2,ws.max_row+1): 
        # if row%100==0:
                # print('Saving Workbook')
                # wb.save(filename = parentPath+sheet +'_'+ str(row)+'_'+fileToProcess)
        if isRefComplete:
            ref     =   str(ws.cell(row,refNoCol).value) 
            batch   =   ref[:len(ref)-12]#str(ws.cell(row,batchCol).value)
            subDiv  =   ref[-12:-7]
            refNo   =   ref[-7:]
        else:
            batch   =   str(ws.cell(row,batchCol).value)
            refNo   =   str(ws.cell(row,refNoCol).value)
        print("Batch : "+str(batch)+"\tSubDiv : "+subDiv+"\tRef No. : "+refNo)

        paramsDict = {
        'nBatchNo':batch,
        'nSubDiv':subDiv,
        'nRefNo':refNo,
        'strRU':'U',
        'submit_param':'submit_value'
        }

        print(str(ws.cell(row,remarksCol).value))

        if str(ws.cell(row,remarksCol).value) != 'None' :
            print("Remarks already present")
            continue
        
        try:
            if batch in [ '24','44','46','27','36']:
                r = requests.post('http://www.lesco.gov.pk:36269/Modules/CustomerBillN/AccountStatusMDI.aspx', data = paramsDict, timeout=10)
            else:
                r = requests.post('http://www.lesco.gov.pk:36269/Modules/CustomerBillN/AccountStatus.aspx', data = paramsDict, timeout=10)
            # r.raise_for_status()
            if r.status_code != requests.codes.ok:
                print(r.status_code)

            soup = BeautifulSoup(r.content, 'html.parser')
            AccountStatus = soup.find(id="ContentPane")    

            k=AccountStatus.find_all('h5')   
            v=AccountStatus.find_all('strong') 
            if batch in [ '24','44','46','27','36']:
                v.pop(0)

            # zip gives a list of tuples of the nth element of each of the lists. However if the list lengths aren't the same, it goes up to the length of the shortest list.
            # zip('foo', '1234') == [('f', '1'), ('o', '2'), ('o', '3')]
            customerInfo = {ke.text.rstrip(':'): va.text for ke, va in zip(k, v)}

            output_txt = 'P='+customerInfo['Amount Paid']
            if output_txt == 'P=0':
                print("No Payment")
                continue
            output_txt += '  DT '+customerInfo['Payment Date']
            output_txt += '  IN '+customerInfo['Bank/Branch']
            ws.cell(row,remarksCol).value = output_txt
            print(output_txt)
        except Exception as e:
            print(e)
            print('No Data Found')

    try:
        print('Completed Sheet '+sheet+' Saving Workbook')
        wb.save(filename = parentPath+'complete_'+sheet+'_'+fileToProcess)
    except:
        pass
print('Completed File Saving Workbook')
try:
    wb.save(filename = parentPath+'complete_'+fileToProcess)
    wb.close()
    
except:
    retry ='r'
    while retry =='r':
        wb.save(filename = parentPath+'complete_'+fileToProcess)
        wb.close()
        retry = input("Enter 'r' to retry, any other key to exit: ")

